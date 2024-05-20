import asyncio
from typing import List

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import joinedload


from bot import get_bot_client
from celery_app.app import celery_app
from core.config import settings
from core.database import AsyncSessionLocal
from model.bot import RCSBot
from model.task import TaskResult
from RCS.google.schema import RCSBatchCapabilityTask, RCSDataForCheck
from utils.excell import ExcellReader


async def save_task_result(task_results):
    async with AsyncSessionLocal() as async_session:
        async_session.add_all(task_results)
        await async_session.commit()


def create_workbook(data, out_file):
    wb = Workbook()

    for sheet_name in data.keys():
        work_sheet = wb.create_sheet(sheet_name)
        for row_num, msisdnd in enumerate(data[sheet_name]):
            work_sheet.cell(row=row_num + 1, column=1).value = msisdnd

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save(out_file)


async def write_task_result_file(task_id):
    async with AsyncSessionLocal() as async_session:
        stmt = select(TaskResult).options(joinedload(TaskResult.task)).where(TaskResult.task_id == task_id)
        results = await async_session.scalars(stmt)
        task_results = results.all()

        task_name = task_results[0].task.name
        data = {}
        for result in task_results:
            if result.country not in data:
                data[result.country] = [result.msisdn]
            else:
                data[result.country].append(result.msisdn)

    out_file = settings.STATIC_PATH.joinpath('tasks', 'results', f'{str(task_name)}.xlsx')
    create_workbook(data, out_file)


@celery_app.task()
def rcs_bulk_check(task_id, task_file_name, bot_name: str):
    task_file = settings.STATIC_PATH.joinpath('tasks').joinpath(task_file_name)
    reader = ExcellReader(task_file)
    data = reader.get_data_for_check()
    task_data = RCSBatchCapabilityTask(task_id=task_id, data=data)
    bot = RCSBot[bot_name]

    loop = asyncio.get_event_loop()
    results = loop.run_until_complete(make_request(bot, task_data.data))
    task_results = []
    for country, msisdns in results.items():
        print(f'{country} with {len(msisdns)}')
        if msisdns:
            task_results.extend([TaskResult(task_id=task_id, country=country, msisdn=msisdn) for msisdn in msisdns])
        else:
            task_results.append(TaskResult(task_id=task_id, country=country))

    loop.run_until_complete(save_task_result(task_results))
    loop.run_until_complete(write_task_result_file(task_id))


async def make_request(bot: RCSBot, data_for_check: List[RCSDataForCheck]):
    coroutines = []
    bot_client = get_bot_client(bot)
    for data in data_for_check:
        coroutine = bot_client.batch_capability(data.msisdns, country=data.country)
        coroutines.append(coroutine)

    try:
        task_result = await asyncio.gather(*[asyncio.create_task(coroutine) for coroutine in coroutines])
        result = await parse_request(task_result)
    finally:
        await bot_client.client.session.close()

    return result


async def parse_request(task_result):
    response = {}
    for result_group in task_result:
        print(f'get responses for {len(result_group)} phones')
        for result in result_group:
            if result.country not in response.keys() and result.rcs_enable:
                response[result.country] = [result.phone_number]
            elif result.country not in response.keys() and not result.rcs_enable:
                response[result.country] = []
            elif result.country in response.keys() and result.rcs_enable:
                response[result.country].append(result.phone_number)

    return response
